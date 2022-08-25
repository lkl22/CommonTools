# -*- coding: utf-8 -*-
# python 3.x
# Filename: OpenpyxlUtil.py
# 定义一个OpenpyxlUtil工具类实现Excel相关的功能 - 2010以上

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from util.LogUtil import *
from util.FileUtil import *


class OpenpyxlUtil:
    @staticmethod
    def getBook(filePath='../resources/mockExam/题库模版.xlsx'):
        """
        获取Book，存在直接读取，不存在创建
        :param filePath:
        :return:
        """
        book = None
        try:
            if FileUtil.existsFile(filePath):
                book = load_workbook(filePath)
            else:
                book = Workbook()
        except Exception as err:
            LogUtil.e('OpenpyxlUtil getBook 错误信息：', err)
        return book

    @staticmethod
    def getSheet(book: Workbook, sheetName='考试信息'):
        """
        获取指定索引的sheet的内容
        :param book: Workbook
        :param sheetName: sheet表的名字
        :return: sheet的内容
        """
        st = None
        try:
            st = book[sheetName]
        except Exception as err:
            LogUtil.e('OpenpyxlUtil getSheet 错误信息：', err)
        return st

    @staticmethod
    def getRows(st: Worksheet):
        """
        获取excel行数
        :param st: sheet
        :return: 行数
        """
        rows = st.max_row
        return rows

    @staticmethod
    def getColumns(st: Worksheet):
        """
        获取excel列数
        :param st: sheet
        :return: 列数
        """
        cols = st.max_column
        return cols

    @staticmethod
    def getCell(st: Worksheet, row, col):
        """
        获取指定单元格内容
        :param st: sheet
        :param row: 行数，从1开始
        :param col: 列数，从1开始
        :return: 单元格内容
        """
        value = st.cell(row, col).value
        return value

    @staticmethod
    def createBook():
        """
        创建Excel工作薄
        :return: Excel工作薄
        """
        return Workbook()

    @staticmethod
    def addSheet(book: Workbook, sheetName, index):
        """
        添加Excel工作表
        :param book: Workbook
        :param sheetName: 表名
        :param index: 表位置
        :return: Worksheet
        """
        return book.create_sheet(sheetName, index)

    @staticmethod
    def writeSheet(st: Worksheet, row, col, value):
        """
        写入数据
        :param st: Worksheet
        :param row: 行数
        :param col: 列数
        :param value: 内容
        """
        st.cell(row, col).value = value

    @staticmethod
    def saveBook(book: Workbook, fileName):
        """
        保存Excel工作薄
        :param book: Workbook
        :param fileName: 文件名
        """
        book.save(fileName)


if __name__ == "__main__":
    bk = OpenpyxlUtil.getBook()
    ws = OpenpyxlUtil.getSheet(bk)
    print(OpenpyxlUtil.getRows(ws))
    print(OpenpyxlUtil.getColumns(ws))
    print(OpenpyxlUtil.getCell(ws, 1, 1))

    bk.close()